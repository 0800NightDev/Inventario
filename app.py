import os
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, Inventory, Transaction
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill

import sys

app = Flask(__name__)
app.config['SECRET_KEY'] = 'dev-secret-key-inventario'

# Pyinstaller Path Resolving para Bases de Datos e Imágenes Persistentes
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

db_url = os.environ.get('DATABASE_URL')
if db_url:
    # Render returns postgres:// which is deprecated in SQLAlchemy 1.4+, need to replace it with postgresql://
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
else:
    db_path = os.path.join(application_path, 'inventario.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'

app.config['UPLOAD_FOLDER'] = os.path.join(application_path, 'static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
@login_required
def index():
    if current_user.role in ['administrador', 'superusuario']:
        return redirect(url_for('dashboard_admin'))
    return redirect(url_for('dashboard_trabajador'))

def check_nsfw(filepath):
    # Función preventiva Mock NSFW.
    # En un entorno de producción o con más capacidades computacionales se usaría un modelo de TensorFlow o API.
    # Como placeholder devolvemos False (Imagen aparentemente limpia y segura).
    return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            if hasattr(user, 'status') and user.status != 'activo':
                flash('Tu cuenta está pendiente de aprobación por el administrador.', 'error')
                return render_template('login.html')
            login_user(user)
            return redirect(url_for('index'))
        flash('Credenciales inválidas, verifica tu usuario o contraseña.', 'error')
        
    return render_template('login.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        whatsapp = request.form.get('whatsapp_number', '').strip()
        foto = request.files.get('foto_carnet')
        
        if not username or not password or not foto or not foto.filename or not whatsapp:
            flash('Debes completar todos los campos (incluyendo el número de WhatsApp) y subir una foto.', 'error')
            return redirect(url_for('registro'))
            
        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya está registrado.', 'error')
            return redirect(url_for('registro'))
            
        if not allowed_file(foto.filename):
            flash('El archivo de foto no es válido. Formatos sugeridos: JPG, PNG.', 'error')
            return redirect(url_for('registro'))
            
        from werkzeug.utils import secure_filename
        import time
        filename = secure_filename(foto.filename)
        filename = f"user_{int(time.time())}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        foto.save(filepath)
        
        # Validación NSFW preventiva (Filtrado de seguridad de cliente)
        if check_nsfw(filepath):
            os.remove(filepath)
            flash('La imagen subida fue marcada como inapropiada por el filtro de seguridad visual. Intenta nuevamente.', 'error')
            return redirect(url_for('registro'))
            
        # El usuario entra como status 'pendiente' hasta que el Admin lo convalide. 
        nuevo_user = User(
            username=username,
            password_hash=generate_password_hash(password, method='pbkdf2:sha256:600000'),
            role='trabajador',
            status='pendiente',
            whatsapp_number=whatsapp,
            foto_carnet=filename
        )
        db.session.add(nuevo_user)
        db.session.commit()
        
        flash('Registro completo. Espera a que un administrador valide tu foto tipo carnet y apruebe tu cuenta.', 'success')
        return redirect(url_for('login'))
        
    return render_template('registro.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/gestion_personal')
@login_required
def gestion_personal():
    if current_user.role not in ['administrador', 'superusuario']:
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard_trabajador'))
    usuarios = User.query.filter(User.role != 'superusuario').all()
    return render_template('gestion_personal.html', usuarios=usuarios)

@app.route('/accion_usuario/<int:user_id>', methods=['POST'])
@login_required
def accion_usuario(user_id):
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    user = User.query.get_or_404(user_id)
    accion = request.form.get('accion')
    
    if user.role == 'superusuario':
        flash('No puedes modificar al superusuario.', 'error')
        return redirect(url_for('gestion_personal'))
        
    if accion == 'aprobar':
        user.status = 'activo'
        db.session.commit()
        flash(f'Usuario {user.username} aprobado y activado.', 'success')
    elif accion == 'eliminar':
        # Primero eliminar la foto 
        if getattr(user, 'foto_carnet', None):
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], user.foto_carnet))
            except:
                pass
        db.session.delete(user)
        db.session.commit()
        flash(f'Usuario {user.username} eliminado Permanentemente.', 'success')
    elif accion == 'reset_password':
        nueva_clave = request.form.get('nueva_clave')
        if nueva_clave and nueva_clave.strip():
            user.password_hash = generate_password_hash(nueva_clave.strip(), method='pbkdf2:sha256:600000')
            db.session.commit()
            flash(f'Contraseña de {user.username} reestablecida exitosamente.', 'success')
        else:
            flash('Error: Debes asignar una contraseña no vacía.', 'error')
            
    elif accion == 'update_whatsapp':
        nuevo_wa = request.form.get('nuevo_whatsapp')
        if nuevo_wa and nuevo_wa.strip():
            user.whatsapp_number = nuevo_wa.strip()
            db.session.commit()
            flash(f'Contacto de WhatsApp de {user.username} modificado a {nuevo_wa}.', 'success')
        else:
            flash('Debes proveer un número de teléfono válido.', 'error')
            
    return redirect(url_for('gestion_personal'))

@app.route('/exportar_excel')
@login_required
def exportar_excel():
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    import io
    import datetime
    
    wb = openpyxl.Workbook()
    
    # Hoja de Mando (Inventario Actual)
    ws_inv = wb.active
    ws_inv.title = "Inventario Dashboard"
    ws_inv.append(["Tamaño Placa", "Formato", "Total Cajas", "Desglose (Bultos y Cajas)"])
    header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
    for cell in ws_inv[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        
    inventario = Inventory.query.all()
    for item in inventario:
        bultos = item.cantidad_cajas // 5
        cajas_sueltas = item.cantidad_cajas % 5
        ws_inv.append([item.tamano_placa, item.formato, item.cantidad_cajas, f"{bultos} bulto(s), {cajas_sueltas} caja(s)"])
    for col in ws_inv.columns:
        ws_inv.column_dimensions[col[0].column_letter].width = 25
        
    # Hoja de Calendario de Ventas/Ingresos Mensuales
    ws_mov = wb.create_sheet(title="Movimientos del Mes")
    ws_mov.append(["ID", "Fecha", "Tipo Operación", "Trabajador", "Tamano Placa", "Formato", "Cajas", "Estado", "Admin Autorizó", "Foto Comprobante"])
    for cell in ws_mov[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        
    mes_actual = datetime.datetime.utcnow().month
    movimientos = Transaction.query.filter(db.extract('month', Transaction.fecha) == mes_actual).order_by(Transaction.fecha.desc()).all()
    
    for idx, tx in enumerate(movimientos, start=2):
        aprobador = tx.admin.username if tx.admin else "Pendiente/Sistema"
        ws_mov.row_dimensions[idx].height = 80
        ws_mov.append([
            tx.id,
            tx.fecha.strftime('%d/%m/%Y %I:%M %p'),
            tx.tipo.upper(),
            tx.user.username,
            tx.tamano_placa,
            tx.formato,
            tx.cantidad_cajas,
            tx.estado.upper(),
            aprobador,
            "" 
        ])
        
        if tx.imagen_path:
            img_path = os.path.join(app.config['UPLOAD_FOLDER'], tx.imagen_path)
            if os.path.exists(img_path):
                try:
                    img = OpenpyxlImage(img_path)
                    img.height = 100
                    img.width = 100
                    ws_mov.add_image(img, f"J{idx}")
                except Exception:
                    pass
                
    for col in ws_mov.columns:
        ws_mov.column_dimensions[col[0].column_letter].width = 20
    ws_mov.column_dimensions["J"].width = 15
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    from flask import send_file
    return send_file(out, download_name=f"Reporte_Ventas_Inventario_{datetime.datetime.utcnow().strftime('%Y_%m')}.xlsx", as_attachment=True)

@app.route('/dashboard_admin')
@login_required
def dashboard_admin():
    if current_user.role not in ['administrador', 'superusuario']:
        flash('Acceso denegado: se requieren permisos de administrador.', 'error')
        return redirect(url_for('dashboard_trabajador'))
    
    transacciones_pendientes = Transaction.query.filter_by(estado='pendiente').order_by(Transaction.fecha.desc()).all()
    
    pendientes_lotes = {}
    for tx in transacciones_pendientes:
        lote_id = tx.imagen_path if tx.imagen_path else f"tx_{tx.id}"
        if lote_id not in pendientes_lotes:
            pendientes_lotes[lote_id] = {
                'fecha': tx.fecha,
                'tipo': tx.tipo,
                'user': tx.user,
                'imagen_path': tx.imagen_path,
                'items': []
            }
        pendientes_lotes[lote_id]['items'].append(tx)
        
    inventario_actual = Inventory.query.all()
    transacciones_aprobadas = Transaction.query.filter_by(estado='aprobada').order_by(Transaction.fecha.desc()).limit(50).all()
    trabajadores = User.query.filter_by(role='trabajador', status='activo').all()
    return render_template('dashboard_admin.html', pendientes_lotes=pendientes_lotes, inventario=inventario_actual, historial=transacciones_aprobadas, trabajadores=trabajadores)

@app.route('/dashboard_trabajador')
@login_required
def dashboard_trabajador():
    pedidos_pendientes = Transaction.query.filter_by(user_id=current_user.id, estado='asignado', tipo='egreso').order_by(Transaction.fecha.desc()).all()
    return render_template('dashboard_trabajador.html', pedidos_pendientes=pedidos_pendientes)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/registro_operacion', methods=['POST'])
@login_required
def registro_operacion():
    tipo = request.form.get('tipo')
    tamanos_seleccionados = request.form.getlist('tamanos')
    fecha_manual_dia = request.form.get('fecha_manual_dia')
    fecha_manual_hh = request.form.get('fecha_manual_hh')
    fecha_manual_mm = request.form.get('fecha_manual_mm')
    fecha_manual_ampm = request.form.get('fecha_manual_ampm')
    
    import datetime
    fecha_operacion = datetime.datetime.utcnow()
    if fecha_manual_dia and fecha_manual_hh and fecha_manual_mm and current_user.role in ['administrador', 'superusuario']:
        try:
            fecha_str = f"{fecha_manual_dia} {fecha_manual_hh}:{fecha_manual_mm} {fecha_manual_ampm}"
            fecha_operacion = datetime.datetime.strptime(fecha_str, '%Y-%m-%d %I:%M %p')
        except ValueError:
            pass
            
    if not tamanos_seleccionados:
        flash('Error: Debes seleccionar al menos un tamaño de placa en tu lote.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    fecha_vencimiento = request.form.get('fecha_vencimiento') or None
    foto = request.files.get('foto')
    
    filename = None
    if foto and foto.filename and allowed_file(foto.filename):
        from werkzeug.utils import secure_filename
        import time
        filename = secure_filename(foto.filename)
        filename = f"tx_{int(time.time())}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        foto.save(filepath)
    elif current_user.role not in ['administrador', 'superusuario']:
        flash('Foto del lote comprobante obligatoria y válida de formato faltante.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    
    
    map_labels = {
        '14x14': '14x14',
        '14x17': '14x17',
        '10x14': '10x14 (26x36)',
        '10x12': '10x12'
    }
    
    operaciones_guardadas = 0
    for tamano in tamanos_seleccionados:
        formatos = request.form.getlist(f'formatos_{tamano}')
        for formato in formatos:
            bult_str = request.form.get(f'bultos_{tamano}_{formato}', '').strip()
            caj_str = request.form.get(f'cajas_{tamano}_{formato}', '').strip()
            
            bultos = int(bult_str) if bult_str else 0
            cajas = int(caj_str) if caj_str else 0
                
            total_cajas = (bultos * 5) + cajas
            if total_cajas > 0:
                t_label = map_labels.get(tamano, tamano)
                
                estado_operacion = 'pendiente'
                admin_aprobador = None
                
                if tipo == 'ingreso':
                    estado_operacion = 'aprobada'
                    admin_aprobador = current_user.id
                    
                    # Afectar inventario directamente
                    inv = Inventory.query.filter_by(tamano_placa=t_label, formato=formato, fecha_vencimiento=fecha_vencimiento).first()
                    if not inv:
                        inv = Inventory(tamano_placa=t_label, formato=formato, cantidad_cajas=total_cajas, fecha_vencimiento=fecha_vencimiento)
                        db.session.add(inv)
                    else:
                        inv.cantidad_cajas += total_cajas

                nueva_tx = Transaction(
                    fecha=fecha_operacion,
                    tipo=tipo,
                    user_id=current_user.id,
                    tamano_placa=t_label,
                    formato=formato,
                    fecha_vencimiento=fecha_vencimiento,
                    cantidad_cajas=total_cajas,
                    imagen_path=filename,
                    estado=estado_operacion,
                    admin_id=admin_aprobador
                )
                db.session.add(nueva_tx)
                operaciones_guardadas += 1
                
    if operaciones_guardadas == 0:
        os.remove(filepath)
        flash('Rechazado: No logramos procesar este ingreso/egreso. Verifica asignar una cantidad mayor que 0 a los subelementos marcados.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    db.session.commit()
    flash(f'Éxito: Se crearon {operaciones_guardadas} solicitud(es) registradas con éxito.', 'success')
    return redirect(url_for('dashboard_trabajador'))

@app.route('/asignar_pedido', methods=['POST'])
@login_required
def asignar_pedido():
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    trabajador_id = request.form.get('trabajador_id')
    if not trabajador_id:
        flash('Error: Debes seleccionar a un trabajador para asignarle el pedido.', 'error')
        return redirect(url_for('dashboard_admin'))
        
    tamanos_seleccionados = request.form.getlist('tamanos_pedido')
    if not tamanos_seleccionados:
        flash('Error: Debes seleccionar tamaños para despachar.', 'error')
        return redirect(url_for('dashboard_admin'))
        
    map_labels = {
        '14x14': '14x14',
        '14x17': '14x17',
        '10x14': '10x14 (26x36)',
        '10x12': '10x12'
    }
    
    operaciones_guardadas = 0
    import datetime
    fecha_operacion = datetime.datetime.utcnow()
    
    for tamano in tamanos_seleccionados:
        formatos = request.form.getlist(f'formatos_pedido_{tamano}')
        for formato in formatos:
            vencimiento = request.form.get(f'vencimiento_pedido_{tamano}_{formato}') or None
            b_str = request.form.get(f'bultos_pedido_{tamano}_{formato}', '').strip()
            c_str = request.form.get(f'cajas_pedido_{tamano}_{formato}', '').strip()
            
            bultos = int(b_str) if b_str else 0
            cajas = int(c_str) if c_str else 0
                
            total_cajas = (bultos * 5) + cajas
            if total_cajas > 0:
                t_label = map_labels.get(tamano, tamano)
                nueva_tx = Transaction(
                    fecha=fecha_operacion,
                    tipo='egreso',
                    user_id=trabajador_id,
                    admin_id=current_user.id,
                    tamano_placa=t_label,
                    formato=formato,
                    fecha_vencimiento=vencimiento,
                    cantidad_cajas=total_cajas,
                    estado='asignado',
                    imagen_path=None
                )
                db.session.add(nueva_tx)
                operaciones_guardadas += 1
                
    if operaciones_guardadas > 0:
        db.session.commit()
        flash(f'Pedido Asignado exitosamente. Se delegaron {operaciones_guardadas} bloque(s) al trabajador.', 'success')
        
        # Trigger Notificaciones Nativas y WhatsApp Web API
        try:
            worker = User.query.get(trabajador_id)
            try:
                from plyer import notification
                notification.notify(
                    title='Pedido Despachado 🚀',
                    message=f'Se asignaron {operaciones_guardadas} bloques a {worker.username} exitosamente.',
                    app_name='Inventario',
                    timeout=5
                )
            except Exception as e:
                print(f"No se pudo mostrar la notificación (entorno en la nube/headless): {e}")
            
            if worker and worker.whatsapp_number:
                import urllib.parse
                import webbrowser
                msg = f"🔔 *Alerta de Despacho Móvil*\nHola {worker.username}, tienes un nuevo pedido de extracción asignado en tu Panel esperando por comprobante fotográfico."
                msg_encoded = urllib.parse.quote(msg)
                wa_url = f"https://wa.me/{worker.whatsapp_number}?text={msg_encoded}"
                try:
                    webbrowser.open(wa_url)
                except Exception as wb_e:
                    print(f"No se pudo abrir el navegador de WhatsApp (entorno en la nube): {wb_e}")
        except Exception as e:
            print(f"Error general en notificaciones: {e}")
        except Exception as e:
            print(f"Error en notificaciones: {e}")
            
    else:
        flash('No se especificaron cantidades válidas para crear el pedido.', 'error')
        
    return redirect(url_for('dashboard_admin'))

@app.route('/validar_transaccion/<int:tx_id>', methods=['POST'])
@login_required
def validar_transaccion(tx_id):
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    accion = request.form.get('accion')
    tx = Transaction.query.get_or_404(tx_id)
    
    if tx.estado != 'pendiente':
        flash('Esta transacción ya fue procesada.', 'error')
        return redirect(url_for('dashboard_admin'))
        
    if accion == 'rechazar':
        tx.estado = 'rechazada'
        tx.admin_id = current_user.id
        db.session.commit()
        flash('Transacción rechazada.', 'success')
    elif accion == 'aprobar':
        inv = Inventory.query.filter_by(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento).first()
        if not inv:
            inv = Inventory(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento, cantidad_cajas=0)
            db.session.add(inv)
            
        if tx.tipo == 'ingreso':
            inv.cantidad_cajas += tx.cantidad_cajas
        elif tx.tipo == 'egreso':
            if inv.cantidad_cajas < tx.cantidad_cajas:
                flash(f'Error: Solo hay {inv.cantidad_cajas} cajas en el lote {tx.fecha_vencimiento} y se intentan egresar {tx.cantidad_cajas}.', 'error')
                return redirect(url_for('dashboard_admin'))
            inv.cantidad_cajas -= tx.cantidad_cajas
            
        tx.estado = 'aprobada'
        tx.admin_id = current_user.id
        db.session.commit()
        flash('Transacción aprobada. Inventario actualizado.', 'success')
        
    return redirect(url_for('dashboard_admin'))

@app.route('/validar_lote/<path:lote_id>', methods=['POST'])
@login_required
def validar_lote(lote_id):
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    accion = request.form.get('accion')
    
    if lote_id.startswith('tx_') and len(lote_id.split('_')) == 2 and lote_id.split('_')[1].isdigit():
        tx_id = int(lote_id.split('_')[1])
        transacciones = Transaction.query.filter_by(id=tx_id, estado='pendiente').all()
    else:
        transacciones = Transaction.query.filter_by(imagen_path=lote_id, estado='pendiente').all()
        
    if not transacciones:
        flash('Lote no encontrado o ya procesado.', 'error')
        return redirect(url_for('dashboard_admin'))
        
    if accion == 'rechazar':
        for tx in transacciones:
            tx.estado = 'rechazada'
            tx.admin_id = current_user.id
        db.session.commit()
        flash('Lote rechazado y descartado exitosamente.', 'success')
    elif accion == 'aprobar':
        # Validar si hay suficiente stock para egresos antes de comprometer cualquier cosa
        for tx in transacciones:
            if tx.tipo == 'egreso':
                inv = Inventory.query.filter_by(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento).first()
                if not inv or inv.cantidad_cajas < tx.cantidad_cajas:
                    stock_real = inv.cantidad_cajas if inv else 0
                    flash(f'Error: Stock insuficiente. Hay {stock_real} cajas de {tx.tamano_placa} {tx.formato} (Venc: {tx.fecha_vencimiento}) y se intentan egresar {tx.cantidad_cajas}. Se abortó toda la orden.', 'error')
                    db.session.rollback()
                    return redirect(url_for('dashboard_admin'))
                    
        # Aplicar los cambios
        for tx in transacciones:
            inv = Inventory.query.filter_by(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento).first()
            if not inv:
                inv = Inventory(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento, cantidad_cajas=0)
                db.session.add(inv)
            
            if tx.tipo == 'ingreso':
                inv.cantidad_cajas += tx.cantidad_cajas
            elif tx.tipo == 'egreso':
                inv.cantidad_cajas -= tx.cantidad_cajas
                
            tx.estado = 'aprobada'
            tx.admin_id = current_user.id
            
        db.session.commit()
        flash(f'Lote aprobado exitosamente. Se procesaron {len(transacciones)} item(s).', 'success')
        
    return redirect(url_for('dashboard_admin'))

@app.route('/eliminar_transaccion/<int:tx_id>', methods=['POST'])
@login_required
def eliminar_transaccion(tx_id):
    if current_user.role not in ['administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    tx = Transaction.query.get_or_404(tx_id)
    
    # Revertir el inventario afectado si la transacción ya estaba aprobada
    if tx.estado == 'aprobada':
        inv = Inventory.query.filter_by(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento).first()
        if inv:
            if tx.tipo == 'ingreso':
                # Si quitamos un ingreso, restamos del inventario
                inv.cantidad_cajas = max(0, inv.cantidad_cajas - tx.cantidad_cajas)
            elif tx.tipo == 'egreso':
                # Si quitamos un egreso, devolvemos al inventario
                inv.cantidad_cajas += tx.cantidad_cajas
                
    # Eliminar archivo físico de imagen si lo tiene
    if tx.imagen_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], tx.imagen_path))
        except:
            pass
            
    db.session.delete(tx)
    db.session.commit()
    flash('Registro de movimiento eliminado correctamente.', 'success')
    return redirect(url_for('dashboard_admin'))

@app.route('/confirmar_pedido/<int:tx_id>', methods=['POST'])
@login_required
def confirmar_pedido(tx_id):
    if current_user.role not in ['trabajador', 'administrador', 'superusuario']:
        return "Acceso denegado", 403
        
    tx = Transaction.query.get_or_404(tx_id)
    if tx.estado != 'asignado':
        flash('Este pedido ya no está en estado asignado.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    foto = request.files.get('foto')
    if not foto or not foto.filename or not allowed_file(foto.filename):
        flash('Es obligatorio adjuntar una foto de comprobante válida.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    # Verificar que el stock sí sea suficiente
    inv = Inventory.query.filter_by(tamano_placa=tx.tamano_placa, formato=tx.formato, fecha_vencimiento=tx.fecha_vencimiento).first()
    if not inv or inv.cantidad_cajas < tx.cantidad_cajas:
        stock_real = inv.cantidad_cajas if inv else 0
        flash(f'Error Fatal: Stock insuficiente actual ({stock_real} cajas en Lote Vencimiento {tx.fecha_vencimiento}). Imposible despachar el pedido.', 'error')
        return redirect(url_for('dashboard_trabajador'))
        
    from werkzeug.utils import secure_filename
    import time
    filename = secure_filename(foto.filename)
    filename = f"tx_asign_{int(time.time())}_{filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    foto.save(filepath)
    
    # Restar de inventario y aprobar
    inv.cantidad_cajas -= tx.cantidad_cajas
    tx.imagen_path = filename
    tx.estado = 'aprobada'
    # Actualizamos la fecha a la real de confirmación
    import datetime
    tx.fecha = datetime.datetime.utcnow()
    
    db.session.commit()
    flash('Pedido despachado y confirmado exitosamente.', 'success')
    return redirect(url_for('dashboard_trabajador'))

@app.route('/generar_datos_prueba', methods=['POST'])
@login_required
def generar_datos_prueba():
    if current_user.role != 'superusuario':
        return "Acceso denegado", 403
        
    import datetime
    import random
    
    tamanos = ['14x14', '14x17', '10x14 (26x36)', '10x12']
    formatos = ['DI-HL', 'HR-U']
    hoy = datetime.datetime.utcnow()
    
    for _ in range(25):
        tipo = random.choices(['ingreso', 'egreso'], weights=[0.7, 0.3])[0]
        tam = random.choice(tamanos)
        formato = random.choice(formatos)
        cantidad = random.randint(5, 40)
        
        # Generar fecha aleatoria dentro del mes actual
        dia = random.randint(1, hoy.day if hoy.day > 1 else 28)
        fecha_fake = hoy.replace(day=dia, hour=random.randint(7, 18), minute=random.randint(0, 59))
        
        inv = Inventory.query.filter_by(tamano_placa=tam, formato=formato).first()
        if not inv:
            inv = Inventory(tamano_placa=tam, formato=formato, cantidad_cajas=0)
            db.session.add(inv)
            
        if tipo == 'ingreso':
            inv.cantidad_cajas += cantidad
        else:
            if inv.cantidad_cajas < cantidad:
                cantidad = inv.cantidad_cajas
            inv.cantidad_cajas -= cantidad
            
        if cantidad > 0:
            tx = Transaction(
                fecha=fecha_fake,
                tipo=tipo,
                estado='aprobada',
                user_id=current_user.id,
                admin_id=current_user.id,
                tamano_placa=tam,
                formato=formato,
                cantidad_cajas=cantidad
            )
            db.session.add(tx)
            
    db.session.commit()
    flash('Se agregaron varios datos de prueba para gráficos (Ingresos/Egresos).', 'success')
    return redirect(url_for('dashboard_admin'))

@app.route('/eliminar_todo', methods=['POST'])
@login_required
def eliminar_todo():
    if current_user.role != 'superusuario':
        return "Acceso denegado", 403
        
    Transaction.query.delete()
    Inventory.query.delete()
    db.session.commit()
    flash('ATENCIÓN: Se ha eliminado completamente todo el inventario y el historial.', 'success')
    return redirect(url_for('dashboard_admin'))

# Se inicializa la BD si no existe y crea el superusuario
def init_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(role='superusuario').first():
            from werkzeug.security import generate_password_hash
            # Credenciales por defecto (desarrollo): admin / admin123
            su = User(username='admin', password_hash=generate_password_hash('admin123'), role='superusuario', status='activo')
            # También agregamos trabajador y admin de prueba
            trab = User(username='trabajador', password_hash=generate_password_hash('trabajador123'), role='trabajador', status='activo')
            adm = User(username='administrador', password_hash=generate_password_hash('admin123'), role='administrador', status='activo')
            
            db.session.add_all([su, trab, adm])
            db.session.commit()
            print("Base de datos inicializada y usuarios por defecto creados.")
        
        # Opcional: Asegurar que todos los administradores/superusuarios existentes estén activos por si acaso
        admins = User.query.filter(User.role.in_(['administrador', 'superusuario'])).all()
        for a in admins:
            if a.status != 'activo':
                a.status = 'activo'
        db.session.commit()

if __name__ == '__main__':
    # Inicializar la base de datos (crear tablas y usuarios por defecto)
    init_db()
    port = int(os.environ.get('PORT', 8000))
    app.run(host='0.0.0.0', port=port, debug=True)
